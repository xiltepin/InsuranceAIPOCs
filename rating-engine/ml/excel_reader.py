"""
excel_reader.py
Parses the Japan actuarial rating manual Excel file and exposes
factor tables as Python dicts for use by the hybrid rating engine.
"""
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(__file__).parent.parent / "data" / "japan_auto_rating_manual.xlsx"


def _load(path=None):
    p = Path(path) if path else EXCEL_PATH
    if not p.exists():
        raise FileNotFoundError(
            f"Rating manual not found at {p}. "
            "Upload japan_auto_rating_manual.xlsx to rating-engine/data/"
        )
    return openpyxl.load_workbook(p, data_only=True)


def get_ncd_factors(path=None):
    wb = _load(path)
    ws = wb["NCD_Grades"]
    factors = {}
    for row in ws.iter_rows(min_row=5, max_row=24, values_only=True):
        grade, bi, pd_, veh, pax = row[1], row[2], row[3], row[4], row[5]
        if grade is not None:
            factors[int(grade)] = {
                "bi": float(bi), "pd": float(pd_),
                "vehicle": float(veh), "passenger": float(pax),
            }
    return factors


AGE_CONDITION_KEYS = ["all", "21+", "26+", "30+", "35+"]

def get_age_factors(path=None):
    wb = _load(path)
    ws = wb["Age_Factors"]
    factors = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=8, values_only=True)):
        cond, bi, pd_, veh, pax = row[1], row[2], row[3], row[4], row[5]
        if cond is not None and i < len(AGE_CONDITION_KEYS):
            factors[AGE_CONDITION_KEYS[i]] = {
                "bi": float(bi), "pd": float(pd_),
                "vehicle": float(veh), "passenger": float(pax),
            }
    return factors


def get_prefecture_factors(path=None):
    wb = _load(path)
    ws = wb["Prefecture_Rates"]
    factors = {}
    for row in ws.iter_rows(min_row=4, max_row=50, values_only=True):
        code, pref, bi_pd, veh, cls = row[1], row[2], row[3], row[4], row[5]
        if code is not None:
            factors[str(code).zfill(2)] = {
                "bi_pd": float(bi_pd), "vehicle": float(veh),
                "region_class": cls,
            }
    return factors


def get_vehicle_factors(path=None):
    wb = _load(path)
    ws = wb["Vehicle_Class"]
    factors = {}
    for row in ws.iter_rows(min_row=4, max_row=20, values_only=True):
        cat, disp, bi_pd, veh, cls = row[1], row[2], row[3], row[4], row[5]
        if cls is not None:
            factors[int(cls)] = {
                "bi_pd": float(bi_pd), "vehicle": float(veh),
                "category": str(cat),
            }
    return factors


DRIVER_RESTRICTION_KEYS = ["none", "family", "spouse", "self"]

def get_driver_restriction_factors(path=None):
    wb = _load(path)
    ws = wb["Driver_Restriction"]
    factors = {}
    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=7, values_only=True)):
        dtype, bi, veh, pax = row[1], row[2], row[3], row[4]
        if dtype is not None and i < len(DRIVER_RESTRICTION_KEYS):
            factors[DRIVER_RESTRICTION_KEYS[i]] = {
                "bi_pd": float(bi), "vehicle": float(veh),
                "passenger": float(pax),
            }
    return factors


def get_base_premiums(path=None):
    wb = _load(path)
    ws = wb["Base_Premiums"]
    coverage_keys = ["bi", "pd", "vehicle", "passenger", "single_car"]
    result = {}
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=9, values_only=True)):
        if i >= len(coverage_keys):
            break
        key = coverage_keys[i]
        result[key] = {}
        for j, cls in enumerate([1, 3, 5, 7, 9, 11]):
            val = row[2 + j]
            if val is not None:
                result[key][cls] = float(val)
    return result


def load_all_factors(path=None):
    return {
        "ncd":                get_ncd_factors(path),
        "age":                get_age_factors(path),
        "prefecture":         get_prefecture_factors(path),
        "vehicle":            get_vehicle_factors(path),
        "driver_restriction": get_driver_restriction_factors(path),
        "base_premiums":      get_base_premiums(path),
    }


def excel_calculate_premium(inputs: dict, factors: dict) -> dict:
    """
    Approach 2 — pure Excel actuarial chain.
    Equivalent to what Drools would execute.
    base_rate x NCD x age x prefecture x driver_restriction
    """
    ncd_grade    = int(inputs.get("ncd_grade", 6))
    age_cond     = inputs.get("age_condition", "26+")
    pref_code    = str(inputs.get("prefecture_code", "13")).zfill(2)
    vehicle_cls  = int(inputs.get("vehicle_rating_class", 5))
    driver_restr = inputs.get("driver_restriction", "none")

    ncd  = factors["ncd"].get(ncd_grade,   factors["ncd"][6])
    age  = factors["age"].get(age_cond,    factors["age"]["26+"])
    pref = factors["prefecture"].get(pref_code, {"bi_pd": 1.0, "vehicle": 1.0})
    dr   = factors["driver_restriction"].get(driver_restr,
           factors["driver_restriction"]["none"])

    available_cls = sorted(factors["base_premiums"]["bi"].keys())
    nearest_cls   = min(available_cls, key=lambda c: abs(c - vehicle_cls))
    bp            = factors["base_premiums"]

    bi_p  = bp["bi"].get(nearest_cls,  38000) * ncd["bi"]       * age["bi"]       * pref["bi_pd"]  * dr["bi_pd"]
    pd_p  = bp["pd"].get(nearest_cls,  31000) * ncd["pd"]       * age["pd"]       * pref["bi_pd"]  * dr["bi_pd"]
    veh_p = bp["vehicle"].get(nearest_cls, 68000) * ncd["vehicle"] * age["vehicle"] * pref["vehicle"]* dr["vehicle"]
    pax_p = bp["passenger"].get(nearest_cls, 11000) * ncd["passenger"] * age["passenger"] * pref["bi_pd"] * dr["bi_pd"]

    total = bi_p + pd_p + veh_p + pax_p
    return {
        "method":              "excel_actuarial",
        "ncd_grade":           ncd_grade,
        "vehicle_class":       nearest_cls,
        "bi_premium":          round(bi_p),
        "pd_premium":          round(pd_p),
        "vehicle_premium":     round(veh_p),
        "passenger_premium":   round(pax_p),
        "annual_premium_jpy":  round(total),
        "monthly_premium_jpy": round(total / 12),
    }
